from urllib import request
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from django.contrib import messages
import requests
import json
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import BLACK,BLUE
import reportlab 
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from xhtml2pdf import pisa
from django.template.loader import get_template

def send_confirmation_email(mail_data):

  try:
    
    subject = mail_data.get('subject')
    email = mail_data.get('email')
    message = mail_data.get('message')
    request = mail_data.get('request')
  
    message_template = render_to_string('home/student_email_template.html',{'message':message})
    email = EmailMessage(subject,message_template, from_email=settings.DEFAULT_FROM_EMAIL,to=[email])
    email.content_subtype = 'html'
    email.send()
    return 'Email sent Successfully'
  except:
    return 'There was an error sending the email to the Student'


def generate_jwt_token(request):
    username = 'Blackhat'
    password = 'qwertytrewq2'
    auth_base_url = 'http://127.0.0.1:8000/'
    url  = '{}{}'.format(auth_base_url,'fetch-auth-token/')
    payload = {'username':username,'password':password}
    headers= {
    'Accept': 'application/json',                         
    'Content-Type': 'application/json; charset=utf-8'
    }
    response = requests.post(url,headers=headers,data=json.dumps(payload))
    if response.status_code == 200:
        data = response.json()['access']
        return data
    else:
        return messages.add_message(request, messages.ERROR, "An error occured fetching the token")
        



def generate_excel_csv(searched_queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = "attached; filename=student_names.csv"
    student_names = csv.writer(response)
    student_names.writerow(['Number','Name','Email','Gender','Course','Phone Number','Reported On'])
    for student in searched_queryset:
        student_names.writerow([student.id,student.name,student.email,student.gender,
        student.course,student.phone_number,student.reported_on])
    return response


def generate_excel_xlsx(searched_queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    """define the file name downloaded"""
    response['Content-Disposition'] = 'attachment; filename=student_name.xlsx'
    workbook = Workbook()
    workbook.remove(workbook.active)
    """define the font used"""
    header_font = Font(name='Calibre',bold=True)
    centered_allignment = Alignment(horizontal='center')
    """define borders and colors"""
    borders = Border(
        bottom=Side(border_style='medium',color='FF000000'),
        top = Side(border_style='medium',color='FF000000'),
        left = Side(border_style='medium',color='FF000000'),
        right = Side(border_style='medium',color='FF000000'),)
    wrapped_alignment = Alignment(vertical='top',wrap_text=True)
    """add column headers and columns sizes"""
    columns = [
        ('Number',10),
        ('Name',35),
        ('Email',25),
        ('Gender',15),
        ('Course',25),
        ('Phone Number',20),
        ('GPA',20),
        ('Reported On',25)
    ]
    worksheet = workbook.create_sheet(
        title='Exported Students',
        index=1
    )
    """add header colors and fill"""
    fill = PatternFill(
        start_color='b3b3b3',
        end_color='b3b3b3',
        fill_type='solid'
    )
    row_num = 1
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = borders
        cell.alignment = centered_allignment
        cell.fill = fill
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
    for student in searched_queryset:
        row_num+=1
        student_number = student.id
        student_name = student.name
        student_email = student.email
        student_gender = student.gender
        student_course = student.course
        student_gpa = student.gpa
        student_phone_number = student.phone_number
        student_reported_on = str(student.reported_on)
        """define cells and cell types allowed"""
        row = [
            (student_number,'Normal'),
            (student_name,'Normal'),
            (student_email,'Normal'),
            (student_gender,'Normal'),
            (student_course,'Normal'),
            (student_phone_number,'Normal'),
            (student_gpa,'Normal'),
            (student_reported_on,'Normal')
        ]
        for col_num, (cell_value, cell_format) in enumerate(row,1):
            cell = worksheet.cell(row=row_num,column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            cell.alignment = wrapped_alignment
    worksheet.freeze_panes = worksheet['A2']
    worksheet.sheet_properties.tabcolor = '00666699'
    workbook.save(response)
    return response
    


 
def generate_reportlab_pdf(request):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 100, "Hello world.")
    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='hello.pdf')



def render_to_pdf(template_src,context_dict):
    template = get_template(template_src)
    html  = template.render(context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode('ISO-8859-1')), result)
    pdf = result.getvalue()
    return pdf


def generate_pdf(request,searched_queryset):
    context_dict = {'searched_queryset':searched_queryset,}
    pdf = render_to_pdf('home/pdf_template.html',context_dict)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        content = "attachment; filename='students.pdf'"
        response['Content-Disposition'] = content
        return response
